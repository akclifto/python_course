import openpyxl as xl

workbook = xl.load_workbook("src/project_excel_automation/students-result.xlsx") # load a workbook
sheet = workbook["Sheet1"] # select a sheet from the workbook
d4 = sheet["D4"] # select a cell in the sheet
print(d4.value)

# better way to get value in cell without specifying name
print(sheet.cell(4,4).value)

#  get subjects in excel sheet and iterate through scores
for row in range(2, sheet.max_row + 1):
    phy = sheet.cell(row, 3).value
    mat = sheet.cell(row, 4).value
    hist = sheet.cell(row, 5).value
    geo = sheet.cell(row, 6).value
    bio = sheet.cell(row, 7).value
    chem = sheet.cell(row, 8).value
    print(phy, mat, hist, geo, bio, chem)

    # save totals marks for each student in the correct cell in the workbook
    total_marks = phy + mat + hist + geo + bio + chem
    total_marks_cell = sheet.cell(row, 9) # col 9 in the "total marks" cell in the workbook
    total_marks_cell.value = total_marks

workbook.save("src/project_excel_automation/students-result-final.xlsx")
print("total marks added, document saved successfully")
